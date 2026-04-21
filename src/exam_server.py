# This file is part of Servidor de Exámenes Exam-Web by Naidel.
#
# Copyright (C) 2024–2026 by Naidel
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
# See the GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""
exam_server.py

Servidor web del sistema de exámenes.

Este módulo gestiona:
- Página principal
- Flujo de preguntas
- Cálculo de estadísticas globales y de sesión
- Persistencia SOLO al pulsar "Siguiente"
"""

import random
import app_meta

from pathlib import Path
from flask import Flask, redirect, render_template_string, request, url_for
from openpyxl import load_workbook
from html import escape
from werkzeug.serving import make_server
from flask import send_from_directory

import os
import sys

from data.base import AnswerResult, AnswerOption
from data.excel_source import ExcelQuestionRepository
from data.access_source import AccessQuestionRepository
from data.sqlite_source import SQLiteQuestionRepository
from data.base import AnswerResult

# ============================================================
# CONFIGURACIÓN
# ============================================================

DEFAULT_DATA_FILE = Path("Exam-Excel.xlsm")
DEFAULT_PORT = 5000
DEFAULT_PROTOCOL = "HTTP"
DEFAULT_CERT_FILE = ""
DEFAULT_KEY_FILE = ""

DATA_FILE = Path("Exam-Excel.xlsm")
SHEET_NAME = "BD"

http_server = None

BASE_DIR = Path(sys.argv[0]).resolve().parent
STATIC_DIR = BASE_DIR / "static"

app = Flask(
    __name__,
    static_folder=str(STATIC_DIR),
    static_url_path="/static"
)

# ============================================================
# FILTROS DE SESIÓN (CONFIGURACIÓN ACTUAL)
# ============================================================

# Filtros de la sesión (valores de los desplegables)
session_filtro_revisados = "Todos"
session_filtro_fallados = "Todos"

# ============================================================
# CONTADORES DE SESIÓN (VOLÁTILES)
# ============================================================

session_correct = 0
session_incorrect = 0
session_initialized = False
session_mode = None
review_flag_current_question = False

session_current_question_id = None
session_current_correct_option = None


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def get_repository(data_file: str):
    ext = Path(data_file).suffix.lower()

    if ext in (".xls", ".xlsx", ".xlsm"):
        return ExcelQuestionRepository(data_file)

    if ext in (".mdb", ".accdb"):
        return AccessQuestionRepository(data_file)

    if ext == ".db":
        return SQLiteQuestionRepository(data_file)

    raise ValueError("Tipo de base de datos no soportado")

def calculate_global_stats():
    with open_repo() as repo:
        return repo.calculate_global_stats()

def render_question_by_id(question_id):
    global session_correct, session_incorrect, session_initialized, session_mode

    global_stats = calculate_global_stats()

    if not session_initialized:
        if session_mode == "continue":
            session_correct = global_stats["correct"]
            session_incorrect = global_stats["incorrect"]
        else:
            session_correct = 0
            session_incorrect = 0
        session_initialized = True

    session_total = session_correct + session_incorrect
    session_percentage = (
        (session_correct / session_total) * 100
        if session_total > 0 else 0
    )

    with open_repo() as repo:
        data = repo.get_question_detail(question_id)

    if data is None:
        return "<h2>Pregunta no encontrada</h2>"

    # =====================================================
    # ✅ BARAJAR OPCIONES DE RESPUESTA (ANTI‑MEMORIA)
    # =====================================================

    question = data["question"]

    # -----------------------------------------------------
    # ✅ BARAJADO CORRECTO DE RESPUESTAS
    #    (letras fijas, contenidos aleatorios)
    # -----------------------------------------------------

    # Construimos lista de contenidos originales
    answers = [
        {
            "text": question.options[0].text,
            "explanation": question.options[0].explanation,
            "is_correct": question.correct_option == "A",
        },
        {
            "text": question.options[1].text,
            "explanation": question.options[1].explanation,
            "is_correct": question.correct_option == "B",
        },
        {
            "text": question.options[2].text,
            "explanation": question.options[2].explanation,
            "is_correct": question.correct_option == "C",
        },
        {
            "text": question.options[3].text,
            "explanation": question.options[3].explanation,
            "is_correct": question.correct_option == "D",
        },
    ]

    # Barajamos SOLO los contenidos
    random.shuffle(answers)

    # Reasignamos a A/B/C/D en orden fijo
    letters = ["A", "B", "C", "D"]
    new_options = []
    new_correct_letter = None

    for letter, ans in zip(letters, answers):
        new_options.append(
            AnswerOption(letter, ans["text"], ans["explanation"])
        )
        if ans["is_correct"]:
            new_correct_letter = letter

    # Aplicamos al objeto Question
    question.options = new_options
    question.correct_option = new_correct_letter
    
    global session_current_question_id, session_current_correct_option

    session_current_question_id = question_id
    session_current_correct_option = new_correct_letter


    return render_template_string(
        QUESTION_PAGE,
        question=question,
        question_vis=data["vis"],
        question_cor=data["cor"],
        review_checked=data["rev"],
        question_ok=data["ok"],
        question_ko=data["ko"],
        global_stats=global_stats,
        session={
            "correct": session_correct,
            "incorrect": session_incorrect,
            "percentage": round(session_percentage, 2),
        },
    )

def open_repo():
    """
    Devuelve el repositorio adecuado (Excel o Access)
    en modo context manager.
    """
    ext = Path(DATA_FILE).suffix.lower()

    if ext in (".xls", ".xlsx", ".xlsm"):
        return ExcelQuestionRepository(DATA_FILE, SHEET_NAME)

    if ext in (".mdb", ".accdb"):
        return AccessQuestionRepository(DATA_FILE)

    if ext == ".db":
        return SQLiteQuestionRepository(DATA_FILE)

    raise ValueError(f"Tipo de base de datos no soportado: {DATA_FILE}")
    
# ============================================================
# PLANTILLAS HTML
# ============================================================

HOME_PAGE = """
<!doctype html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <title>Examen - Página principal</title>

    <style>
    body {
        font-family: Arial, sans-serif;
        max-width: 700px;
        margin: 40px auto;
        line-height: 1.5;

        /* Fondo acorde al banner */
        background: linear-gradient(
            135deg,
            #d9f3f8 0%,
            #cbeef6 50%,
            #dfeff4 100%
        );

        h1 {
            text-align: center;
        }

        .actions {
            margin: 30px 0;
        }

        .actions button {
            margin-bottom: 10px;
            display: block;
            width: 100%;
        }

        button {
            padding: 10px 16px;
            font-size: 14px;
            cursor: pointer;
        }

        .filters {
            margin: 25px 0;
        }

        select {
            margin-left: 10px;
            padding: 4px;
        }

        .links a {
            display: inline-block;
            margin-right: 12px;
            margin-top: 8px;
        }

        .help {
            margin-top: 40px;
            font-size: 0.9em;
            color: #333;
        }
    </style>
</head>

<body>

<h1>{{ app_name }}</h1>
<div style="text-align:center; margin: 20px 0;">
    <img
      src="{{ url_for('static', filename='exam-web-banner-lite.png') }}"
      alt="Exam-Web - Training system for exams built for the web"
      style="max-width:100%; height:auto; border-radius:12px;"
      onerror="this.style.display='none';"
    >
</div>
<div style="text-align:center; font-size: 0.9em;">
<p>Versión v{{ app_version }} | Build {{ app_build }} ({{ app_estado }})</p>
</div>
<form method="post">

    <div class="filters">
        <label>
            Revisados:
            <select name="filtro_revisados">
                <option value="NoRevisados">Sólo no revisados</option>
                <option value="Todos">Todos</option>
            </select>
        </label>

        <br><br>

        <label>
            Fallados:
            <select name="filtro_fallados">
                <option value="SoloFallados">No incluir los acertados</option>
                <option value="Todos">Todos</option>
            </select>
        </label>
    </div>

    <div class="actions">
        <button type="submit" formaction="{{ url_for('start_exam') }}">
            Comenzar
        </button>

        <button type="submit" formaction="{{ url_for('continue_exam') }}">
            Continuar la sesión
        </button>
    </div>

</form>

<form method="post"
      action="{{ url_for('reset_db') }}"
      onsubmit="return confirm('¿Seguro que deseas resetear la base de datos?');">

    <div class="actions">
        <button type="submit">Resetear Base de Datos</button>
    </div>

</form>

<div class="links">
    <a href="{{ url_for('about') }}" target="_blank">Acerca de...</a>
    <a href="{{ url_for('license') }}" target="_blank">Licencia</a>
    <a href="{{ url_for('instructions') }}" target="_blank">Instrucciones</a>
    <a href="https://paypal.me/atmarquez" target="_blank">❤️ Donar con PayPal</a>
</div>

<div class="help">
    <h3>¿Qué hace cada opción?</h3>

    <p><strong>Comenzar</strong>: inicia una nueva sesión de práctica desde cero,
    respetando los filtros seleccionados.</p>

    <p><strong>Continuar la sesión</strong>: continúa la práctica partiendo
    del estado actual de la base de datos, respetando los filtros.</p>

    <p><strong>Resetear Base de Datos</strong>: borra las estadísticas almacenadas
    en el Excel (VIS, COR, REV, OK, KO).</p>

    <p><strong>Revisados</strong>: controla si se incluyen o excluyen preguntas
    marcadas para revisión.</p>

    <p><strong>Fallados</strong>: controla si se incluyen preguntas ya acertadas.</p>
</div>

</body>
</html>
"""

QUESTION_PAGE = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Pregunta</title>

<style>
body { 
    font-family: Arial, sans-serif; 
    display: flex; 
}
.main { 
    width: 75%; 
    padding: 10px; 
}
.side { 
    width: 25%; 
    background: #e6f6ff; 
    padding: 10px; 
}
.option { 
    cursor: pointer; 
    margin: 6px 0; 
    border: 1px solid #333; 
    padding: 6px; 
}
.correct { 
    color: green; 
}
.incorrect { 
    color: red; 
}

button {
    padding: 2px 6px;
    font-size: 12px;
    margin-left: 4px;
    width: auto;
}

@keyframes blink {
    0%   { opacity: 1; }
    50%  { opacity: 0; }
    100% { opacity: 1; }
}

.timer-normal {
    color: #003366;
    font-weight: normal;
    animation: none;
}

.timer-warning {
    color: #cc5500;
    font-weight: normal;
    animation: blink 1s infinite;
}

.timer-danger {
    color: #8B0000; /* rojo sangre */
    font-weight: bold;
    animation: blink 0.7s infinite;
}

</style>
</head>

<body>

<div class="main">
    <p>
        <div style="margin: 10px 0;">
            <strong>⌚ Sesión:</strong>
            <span id="sessionTimer">00:00</span>

            <strong>⏱️ Pregunta:</strong>
            <span id="questionTimer" style="color:#003366;">00:00</span>

            <button type="button" onclick="playTimers()">▶️</button>
            <button type="button" onclick="pauseTimers()">⏸️</button>
            <button type="button" onclick="resetTimers()">🔄</button>
        </div>
        <strong>TIPO:</strong> {{ question.question_type }} |
        <strong>PARTE:</strong> {{ question.id }} |
        <strong>VISTO:</strong> {{ "SI" if question_vis == 1 else "NO" }} |
        <strong>FALLADO:</strong> {{ "SI" if question_vis == 1 and question_cor == 0 else "NO" }} |
        <strong>OK:</strong> {{ question_ok }} |
        <strong>KO:</strong> {{ question_ko }}
    </p>

    <h2>{{ question.statement }}</h2>
    <div style="margin: 10px 0; font-size: 14px;">
        <a href="#" onclick="openSheet('Gráficos')">Gráficos</a>
        &nbsp;&nbsp;&nbsp;
        <a href="#" onclick="openSheet('Glosario')">Glosario</a>
        &nbsp;&nbsp;&nbsp;
        <a href="#" onclick="openSheet('Anotaciones')">Anotaciones</a>
    </div>
    <div id="options">
    {% for opt in question.options %}
        <div class="option"
             data-key="{{ opt.key }}"
             data-explanation="{{ opt.explanation }}">
            <strong>{{ opt.key }}</strong> {{ opt.text }}
        </div>
    {% endfor %}
    </div>

    <div id="feedback"></div>

    <button id="nextBtn" disabled>Siguiente &gt;&gt;&gt;</button>

    {% if question.study_notes and question.study_notes.strip() %}
    <div id="studyBlock" style="display: none;">
        <h3>Estudio</h3>
        <p>{{ question.study_notes }}</p>
    </div>
    {% endif %}
</div>

<div class="side">
    <h3>Preguntas</h3>
    <p>Total: {{ global_stats.total_questions }}</p>
    <p>Vistas: {{ global_stats.seen_questions }}</p>
    <p>% Vistas: {{ global_stats.seen_percentage }}%</p>

    <hr>
    <h3>Sesión actual</h3>
    <p>Correctas: {{ session.correct }}</p>
    <p>Incorrectas: {{ session.incorrect }}</p>
    <p>% Aciertos: {{ session.percentage }}%</p>

    <h3>Datos globales</h3>
    <p>Correctas: {{ global_stats.correct }}</p>
    <p>Incorrectas: {{ global_stats.incorrect }}</p>
    <p>% Aciertos: {{ global_stats.percentage }}%</p>

    <hr>

    <label>
        <input type="checkbox"
               id="reviewCheckbox"
               {% if review_checked %}checked{% endif %}>
        Marcar para revisión
    </label>
    <br><br>
<button onclick="openDatabase()">🗃️ Ver en la B.D.</button>
<button onclick="refreshQuestion()">🗘 Refrescar pregunta</button>
<button onclick="openCalculator()">🧮 Calculadora</button>
<button onclick="goHome()">🚪 Salir</button
</div>

<script>
/* ================= CRONÓMETROS ================= */

let sessionSeconds = 0;
let questionSeconds = 0;

let sessionInterval = null;
let questionInterval = null;

let timersRunning = false;   // ▶️ en marcha
let timersPaused = false;    // ⏸️ pausado

function formatTime(sec) {
    const m = Math.floor(sec / 60);
    const s = sec % 60;
    return m.toString().padStart(2,'0') + ':' + s.toString().padStart(2,'0');
}

function updateSessionUI() {
    document.getElementById("sessionTimer").textContent =
        formatTime(sessionSeconds);
}

function updateQuestionUI() {
    const qt = document.getElementById("questionTimer");
    qt.textContent = formatTime(questionSeconds);

    // 🔥 ELIMINAR estilos inline que bloquean el color
    qt.style.color = "";
    qt.style.fontWeight = "";

    qt.classList.remove(
        "timer-normal",
        "timer-warning",
        "timer-danger"
    );

    if (questionSeconds >= 90) {
        qt.classList.add("timer-danger");
    } else if (questionSeconds >= 60) {
        qt.classList.add("timer-warning");
    } else {
        qt.classList.add("timer-normal");
    }
}

/* ▶️ */

function playTimers() {
    if (timersRunning) return;

    timersPaused = false;
    timersRunning = true;
    sessionStorage.setItem("timersRunning", "true");

    sessionInterval = setInterval(() => {
        sessionSeconds++;
        sessionStorage.setItem("sessionSeconds", sessionSeconds);
        updateSessionUI();
    }, 1000);

    questionInterval = setInterval(() => {
        questionSeconds++;
        updateQuestionUI();
    }, 1000);
}

/* ⏸️ */

function pauseTimers() {
    timersPaused = true;
    timersRunning = false;
    sessionStorage.setItem("timersRunning", "false");

    if (sessionInterval) clearInterval(sessionInterval);
    if (questionInterval) clearInterval(questionInterval);

    sessionInterval = null;
    questionInterval = null;
}

/* 🔄 */

function resetTimers() {
    pauseTimers();

    sessionSeconds = 0;
    questionSeconds = 0;

    sessionStorage.setItem("sessionSeconds", "0");
    sessionStorage.setItem("timersRunning", "false");

    updateSessionUI();
    updateQuestionUI();
}

function onQuestionLoaded() {
    // Restaurar tiempo de sesión
    const storedSession = sessionStorage.getItem("sessionSeconds");
    if (storedSession !== null) {
        sessionSeconds = parseInt(storedSession, 10);
        updateSessionUI();
    }

    // El tiempo de pregunta SIEMPRE empieza en 0
    questionSeconds = 0;
    updateQuestionUI();

    // Restaurar si estaban corriendo o no
    const wasRunning = sessionStorage.getItem("timersRunning") === "true";

    if (wasRunning) {
        playTimers();   // sesión continúa, pregunta empieza desde 0
    }
}

/* Llamar siempre al final */
onQuestionLoaded();

/* ===== LÓGICA DE LA PREGUNTA ===== */
let selectedOption = null;
let reviewChecked = {{ "true" if review_checked else "false" }};
const correctOption = "{{ question.correct_option }}";
const questionId = "{{ question.id }}";

/* ==== SELECCIÓN VISUAL ==== */
function selectOption(optionElement) {
    // Limpiar selección previa
    document.querySelectorAll(".option").forEach(o => {
        o.style.backgroundColor = "";
        o.style.border = "1px solid #333";
    });

    optionElement.style.backgroundColor = "#d0eaff";
    selectedOption = optionElement.dataset.key;
}

/* ==== CONFIRMAR RESPUESTA ==== */
function confirmAnswer(optionElement) {
    if (!selectedOption || selectedOption !== optionElement.dataset.key) {
        return;
    }

    const feedback = document.getElementById("feedback");
    const explanation = optionElement.dataset.explanation || "";

    // Limpiar estilos previos
    document.querySelectorAll(".option").forEach(o => {
        o.style.border = "1px solid #333";
        o.style.backgroundColor = "";
    });

    if (selectedOption === correctOption) {
        feedback.textContent = "✅ Correcto — " + explanation;
        feedback.className = "correct";

        optionElement.style.backgroundColor = "#d4edda";
        optionElement.style.border = "2px solid #28a745";
    } else {
        feedback.textContent = "❌ Incorrecto — " + explanation;
        feedback.className = "incorrect";

        optionElement.style.backgroundColor = "#f8d7da";
        optionElement.style.border = "2px solid #dc3545";

        // Resaltar la opción correcta
        document.querySelectorAll(".option").forEach(o => {
            if (o.dataset.key === correctOption) {
                o.style.backgroundColor = "#d4edda";
                o.style.border = "2px solid #28a745";
            }
        });
    }

    // Mostrar ESTUDIO si existe
    const studyBlock = document.getElementById("studyBlock");
    if (studyBlock) {
        studyBlock.style.display = "block";
    }
  
    // Activar botón Siguiente
    document.getElementById("nextBtn").disabled = false;
}

/* ==== RATÓN ==== */
document.querySelectorAll(".option").forEach(opt => {

    opt.addEventListener("click", () => {
        selectOption(opt);
    });

    opt.addEventListener("dblclick", () => {
        confirmAnswer(opt);
    });
});

/* ==== TECLADO ==== */
document.addEventListener("keydown", (event) => {
    const key = event.key.toUpperCase();

    const option = document.querySelector(
        `.option[data-key="${key}"]`
    );

    if (option) {
        selectOption(option);
        event.preventDefault();
    }

    if (event.key === "Enter" && selectedOption) {
        const selectedElement = document.querySelector(
            `.option[data-key="${selectedOption}"]`
        );
        if (selectedElement) {
            confirmAnswer(selectedElement);
            event.preventDefault();
        }
    }
});

document.getElementById("reviewCheckbox")
    .addEventListener("change", function () {
        reviewChecked = this.checked;
    });

document.getElementById("nextBtn")
    .addEventListener("click", async () => {
        await fetch("/next", {
            method: "POST",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify({
                question_id: questionId,
                selected_option: selectedOption,
                review_flag: reviewChecked
            })
        });
        window.location.href = "/question";
    });

/* ===== CALCULADORA (VENTANA NUEVA) ===== */
function openCalculator() {
    window.open(
        "/calculator",
        "Calculadora",
        "width=380,height=560,resizable=yes"
    );
}

/* ===== 🗃️ Ver en la B.D. (VENTANA NUEVA) ===== */
function openDatabase() {
    window.open(
        "/question_db/{{ question.id }}",
        "BaseDeDatos",
        "width=700,height=800,resizable=yes"
    );
}

/* ===== Refrescar la pregunta (LEEMOS DE NUEVO DE LA B.D. y limpiamos la respuesta) ===== */
function refreshQuestion() {
    window.location.href = "/question_refresh/{{ question.id }}";
}

function openSheet(name) {
    window.open(
        "/data_view/" + encodeURIComponent(name),
        name,
        "width=900,height=700,resizable=yes"
    );
}

function goHome() {
    window.location.href = "/";
}

</script>

</body>
</html>
"""

CALCULATOR_PAGE = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Calculadora</title>

<style>
body {
    font-family: Arial, sans-serif;
    background: #f4f4f4;
}

.calc {
    width: 320px;
    margin: 40px auto;
    background: #fff;
    padding: 15px;
    border-radius: 8px;
    box-shadow: 0 0 10px rgba(0,0,0,0.2);
}

#display {
    width: 100%;
    height: 45px;
    font-size: 20px;
    text-align: right;
    margin-bottom: 10px;
    padding-right: 5px;
}

.grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 8px;
}

button {
    height: 40px;
    font-size: 16px;
    cursor: pointer;
}
</style>
</head>

<body>

<div class="calc">
    <input id="display" type="text">

    <div class="grid">
        <button onclick="add('7')">7</button>
        <button onclick="add('8')">8</button>
        <button onclick="add('9')">9</button>
        <button onclick="add('/')">÷</button>

        <button onclick="add('4')">4</button>
        <button onclick="add('5')">5</button>
        <button onclick="add('6')">6</button>
        <button onclick="add('*')">×</button>

        <button onclick="add('1')">1</button>
        <button onclick="add('2')">2</button>
        <button onclick="add('3')">3</button>
        <button onclick="add('-')">−</button>

        <button onclick="add('0')">0</button>
        <button onclick="add('.')">.</button>
        <button onclick="result()">=</button>
        <button onclick="add('+')">+</button>

        <button onclick="sqrt()">√</button>
        <button onclick="power2()">x²</button>
        <button onclick="add('**')">xʸ</button>
        <button onclick="clearDisplay()">C</button>
    </div>
</div>

<script>
function add(value) {
    document.getElementById('display').value += value;
}

function clearDisplay() {
    document.getElementById('display').value = "";
}

function result() {
    try {
        document.getElementById('display').value =
            eval(document.getElementById('display').value);
    } catch {
        document.getElementById('display').value = "Error";
    }
}

function sqrt() {
    try {
        let v = eval(document.getElementById('display').value);
        document.getElementById('display').value = Math.sqrt(v);
    } catch {
        document.getElementById('display').value = "Error";
    }
}

function power2() {
    try {
        let v = eval(document.getElementById('display').value);
        document.getElementById('display').value = Math.pow(v, 2);
    } catch {
        document.getElementById('display').value = "Error";
    }
}
</script>

</body>
</html>
"""

QUESTION_DB_PAGE = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Editar pregunta</title>

<style>
body {
    font-family: Arial, sans-serif;
    padding: 20px;
}

label {
    font-weight: bold;
}

input, textarea {
    width: 100%;
    margin-bottom: 10px;
}

textarea {
    height: 80px;
}

button {
    padding: 8px 12px;
}
</style>
</head>

<body>

<h2>Editar pregunta en la B.D.</h2>

<form method="post">
{% for key, value in data.items() %}
    <label>{{ key }}</label>

    {% if key in ["VIS", "COR", "REV"] %}
        <input type="number"
               name="{{ key }}"
               value="{{ value }}"
               min="0"
               max="1"
               step="1">

    {% elif key in ["OK", "KO"] %}
        <input type="number"
               name="{{ key }}"
               value="{{ value }}"
               min="0"
               step="1">

    {% elif key in ["PREGUNTA", "ESTUDIO", "RA", "RB", "RC", "RD"] %}
        <textarea name="{{ key }}">{{ value }}</textarea>

    {% else %}
        <input type="text" name="{{ key }}" value="{{ value }}">
    {% endif %}
{% endfor %}

<button type="submit">Guardar</button>
</form>

</body>
</html>
"""

ABOUT_PAGE = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Acerca de… | Exam‑Web</title>

<style>
body {
    font-family: Arial, sans-serif;
    max-width: 900px;
    margin: 30px auto;
    padding: 0 20px;
    line-height: 1.6;
    color: #222;
}

h1, h2 {
    color: #004b87;
}

.section {
    margin-bottom: 30px;
}

.section h2 {
    border-bottom: 2px solid #e6f6ff;
    padding-bottom: 4px;
}

ul {
    margin-left: 20px;
}

.icon {
    font-size: 1.2em;
    margin-right: 6px;
}

.footer {
    margin-top: 40px;
    font-size: 0.9em;
    color: #555;
}
a {
    color: #0066cc;
    text-decoration: none;
}
a:hover {
    text-decoration: underline;
}
</style>
</head>

<body>

<h1>ℹ️ Acerca de…</h1>

<p><strong>{{ app_name }}</strong><br>
Sistema de entrenamiento para exámenes vía web</p>

<div class="section">
<h2>🔖 Información general</h2>
<p><strong>Nombre del proyecto:</strong> Exam‑Web by Naidel</p>
<p><strong>Descripción:</strong><br>
Exam‑Web es un sistema de entrenamiento y autoevaluación para exámenes tipo test desarrollado íntegramente en Python.
Permite practicar preguntas, simular exámenes reales, analizar fallos y reforzar el estudio mediante estadísticas y retroalimentación detallada.
Está diseñado para certificaciones, oposiciones y estudio técnico, priorizando la reutilización, la flexibilidad y la transparencia del código.
</p>
</div>

<div class="section">
<h2>🏷️ Versión</h2>
<ul>
<li><strong>Versión:</strong> v{{ app_version }}</li>
<li><strong>Estado:</strong> {{ app_estado }}</li>
<li><strong>Fecha de build:</strong> {{ app_build }}</li>
<li><strong>Plataforma:</strong> Microsoft Windows (Desktop)</li>
</ul>
</div>

<div class="section">
<h2>👤 Autor</h2>
<ul>
<li><strong>Autor:</strong> Antonio Teodomiro Márquez Muñoz</li>
<li><strong>Alias:</strong> Naidel</li>
<li><strong>Contacto:</strong> 📧 <a href="mailto:atmarquez@gmail.com">atmarquez@gmail.com</a></li>
</ul>
<p>
Exam‑Web es software libre.  
Si te resulta útil, puedes apoyar su desarrollo con una donación voluntaria.
</p>
<p>❤️ <a href="https://www.paypal.com/donate" target="_blank">Donar con PayPal</a></p>
</div>

<div class="section">
<h2>© Copyright</h2>
<p>Copyright © 2024‑2026<br>
Antonio Teodomiro Márquez Muñoz (Naidel)</p>
</div>

<div class="section">
<h2>⚖️ Licencia</h2>
<p>
Este programa es software libre, distribuido bajo los términos de la
<strong>GNU General Public License versión 3 (GPLv3)</strong>.
</p>

<ul>
<li>✅ Usarlo para cualquier propósito</li>
<li>✅ Estudiar cómo funciona</li>
<li>✅ Modificarlo</li>
<li>✅ Redistribuirlo</li>
</ul>

<p>
Siempre que respetes los términos de dicha licencia.
</p>

<p>
Este programa se distribuye con la esperanza de que sea útil, pero
<strong>SIN NINGUNA GARANTÍA</strong>; sin siquiera la garantía implícita de
COMERCIABILIDAD o IDONEIDAD PARA UN PROPÓSITO PARTICULAR.
</p>
</div>

<div class="section">
<h2>🧩 Componentes principales del sistema</h2>
<ul>
<li><strong>Hoja Principal:</strong> Inicio de sesión, filtros y control del examen</li>
<li><strong>Hoja Pregunta:</strong> Interfaz principal de preguntas y respuestas</li>
<li><strong>Base de Datos (Tabla 'BD' del repositorio):</strong> Repositorio de preguntas, respuestas y estadísticas</li>
<li><strong>Glosario (Tabla 'Glosario' del repositorio):</strong> Términos técnicos y siglas</li>
<li><strong>Gráficos / Figuras (Tabla 'Gráficos' del repositorio):</strong> Imágenes y elementos visuales referenciados</li>
<li><strong>Anotaciones (Tabla 'Anotaciones' del repositorio):</strong> Anotaciones del usuario</li>
</ul>
</div>

<div class="section">
<h2>💡 Filosofía del proyecto</h2>
<p>
Exam‑Web nace como una herramienta personal utilizada con éxito en:
</p>
<ul>
<li>Certificaciones técnicas (CIA, COSO, CISA, CDPSE, CISM, CRISC)</li>
<li>Procesos de oposición</li>
<li>Estudio autodirigido</li>
</ul>

<p>
El objetivo de liberar el proyecto bajo GPLv3 es:
</p>
<ul>
<li>Compartir una herramienta útil</li>
<li>Facilitar su adaptación a distintos exámenes</li>
<li>Permitir que otros la estudien y la mejoren</li>
<li>Mantener un enfoque honesto y transparente</li>
</ul>
</div>

<div class="section">
<h2>✅ Agradecimientos</h2>
<p>
Gracias a todas las comunidades de software libre y a las personas que comparten
conocimiento, experiencias y herramientas que hacen posible proyectos como este.
</p>
</div>

<div class="footer">
<p>Exam‑Web by Naidel · GPLv3</p>
</div>

</body>
</html>
"""

INSTRUCTIONS_PAGE = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Instrucciones | Exam‑Web</title>

<style>
body {
    font-family: Arial, sans-serif;
    max-width: 900px;
    margin: 30px auto;
    padding: 0 20px;
    line-height: 1.6;
    color: #222;
}

h1, h2 {
    color: #004b87;
}

h2 {
    border-bottom: 2px solid #e6f6ff;
    padding-bottom: 4px;
    margin-top: 30px;
}

.section {
    margin-bottom: 30px;
}

ul {
    margin-left: 20px;
}

.footer {
    margin-top: 40px;
    font-size: 0.9em;
    color: #555;
}
</style>
</head>

<body>

<h1>📘 Manual de Usuario</h1>

<p>
<strong>Sistema de Exámenes Web con Excel / Access / DB</strong><br>
Autor: <strong>Naidel</strong>
</p>

<div class="section">
<h2>1️⃣ Descripción general</h2>
<p>
Este sistema Web es un sistema interactivo de entrenamiento para exámenes,
diseñado para:
</p>
<ul>
<li>Practicar preguntas tipo test</li>
<li>Simular exámenes reales</li>
<li>Registrar aciertos y fallos</li>
<li>Facilitar el estudio posterior</li>
</ul>

<p>
El sistema funciona de forma mixta con <strong>Flask</strong> como frontend y una
base de datos basada en <strong>Excel</strong>, <strong>Access</strong> o <strong>SQLite</strong>.
Está pensado para ser configurable, reutilizable y ampliable.
</p>
</div>

<div class="section">
<h2>2️⃣ Estructura de la web</h2>

<h3>📄 Principal</h3>
<p>Pantalla inicial del sistema. Desde aquí se puede:</p>
<ul>
<li>Iniciar una nueva sesión</li>
<li>Continuar una sesión anterior</li>
<li>Resetear estadísticas de la base de datos</li>
</ul>

<h3>📄 Pregunta</h3>
<p>Pantalla principal de examen. Muestra:</p>
<ul>
<li>Enunciado de la pregunta</li>
<li>Respuestas posibles</li>
<li>Indicadores de estado</li>
<li>Estadísticas de la pregunta actual</li>
</ul>
<p><strong>✅ La respuesta se selecciona haciendo doble clic sobre la opción deseada.</strong></p>

<h3>📄 BD (Base de Datos)</h3>
<p>Contiene todas las preguntas del examen, incluyendo:</p>
<ul>
<li>Enunciado</li>
<li>Respuestas A / B / C / D</li>
<li>Respuesta correcta</li>
<li>Retroalimentación</li>
<li>Estadísticas por pregunta</li>
</ul>
<p>⚠️ La base de datos puede ser Excel, Access o SQLite.</p>

<h3>📄 Glosario</h3>
<p>Contiene términos técnicos o siglas usadas en las preguntas.</p>
<ul>
<li>Concepto</li>
<li>Significado</li>
<li>Método de cálculo (si aplica)</li>
</ul>

<h3>📄 Gráficos</h3>
<p>
Incluye imágenes, gráficos o figuras referenciadas en las preguntas mediante
textos como <em>“VER FIGURA 1”</em>.
</p>
</div>

<div class="section">
<h2>3️⃣ Funcionamiento básico</h2>

<h3>▶ Iniciar un examen</h3>
<ol>
<li>Ir a la página Principal</li>
<li>Pulsar el botón <strong>Comenzar</strong></li>
<li>Se carga la primera pregunta aleatoria</li>
</ol>

<h3>✅ Responder una pregunta</h3>
<ul>
<li>Hacer doble clic sobre la respuesta deseada</li>
<li>El sistema:</li>
<ul>
<li>Marca la respuesta en verde o rojo</li>
<li>Muestra la retroalimentación</li>
<li>Guarda la selección</li>
</ul>
</ul>

<h3>⏭ Avanzar a la siguiente pregunta</h3>
<ul>
<li>Pulsar <strong>Siguiente</strong></li>
<li>El sistema:</li>
<ul>
<li>Contabiliza la pregunta anterior</li>
<li>Selecciona una nueva pregunta</li>
<li>Actualiza las estadísticas</li>
</ul>
</ul>

<h3>🔁 Continuar una sesión</h3>
<p>
Desde Principal, pulsar <strong>Continuar sesión</strong>.  
Se regresa a la última pregunta mostrada.
</p>
</div>

<div class="section">
<h2>4️⃣ Indicadores en pantalla (Página Pregunta)</h2>
<ul>
<li><strong>Visto:</strong> indica si la pregunta ya ha sido mostrada</li>
<li><strong>Fallado:</strong> indica si alguna vez se respondió mal</li>
<li><strong>OK / KO:</strong> estadísticas históricas de esa pregunta</li>
<li><strong>Total sesión:</strong> preguntas respondidas en la sesión actual</li>
<li><strong>Porcentaje:</strong> aciertos / total de la sesión</li>
</ul>
</div>

<div class="section">
<h2>5️⃣ Revisión y estudio</h2>
<ul>
<li>Las preguntas pueden marcarse para revisión</li>
<li>Las preguntas marcadas tienen prioridad en la selección aleatoria</li>
<li>Permiten reforzar puntos débiles</li>
<li>La información de estudio se muestra en un área específica de la pregunta</li>
</ul>
</div>

<div class="section">
<h2>6️⃣ Navegación avanzada</h2>

<h3>🔍 Ver pregunta en la Base de Datos</h3>
<p>
Permite abrir la fila correspondiente para:
</p>
<ul>
<li>Revisar el contenido</li>
<li>Editar la pregunta</li>
<li>Ajustar la retroalimentación</li>
</ul>

<h3>🔄 Refrescar pregunta</h3>
<p>
Permite volver a cargar la pregunta actual tras modificarla en la BD, sin alterar
sus estadísticas.
</p>
</div>

<div class="section">
<h2>7️⃣ Aleatoriedad del sistema</h2>
<p>
El sistema utiliza varias capas de aleatoriedad:
</p>
<ul>
<li>Selección aleatoria de preguntas</li>
<li>Reordenación aleatoria de respuestas</li>
<li>Prioridad a preguntas marcadas en revisión</li>
<li>Respeto de filtros activos</li>
</ul>
<p>
Esto evita memorizar posiciones, sesgos en el orden y repeticiones innecesarias.
</p>
</div>

<div class="section">
<h2>8️⃣ Uso recomendado</h2>
<p>
Este sistema está pensado para:
</p>
<ul>
<li>Preparación de certificaciones</li>
<li>Oposiciones</li>
<li>Estudios técnicos</li>
<li>Autoevaluación continua</li>
</ul>
</div>

<div class="section">
<h2>9️⃣ Nota final</h2>
<p>
Este sistema puede compartirse, modificarse y ampliarse, siempre respetando:
</p>
<ul>
<li>La estructura de hojas</li>
<li>Los encabezados de la base de datos</li>
<li>El funcionamiento interno de la sesión</li>
</ul>
<p><strong>✅ Fin del manual</strong></p>
</div>

<div class="footer">
<p>Exam‑Web by Naidel · Manual de Usuario</p>
</div>

</body>
</html>
"""

# ============================================================
# RUTAS
# ============================================================

@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory(STATIC_DIR, filename)
    
@app.route("/calculator")
def calculator():
    """
    Página de calculadora avanzada independiente.
    """
    return render_template_string(CALCULATOR_PAGE)

@app.route("/question_refresh/<question_id>", methods=["GET"])
def question_refresh(question_id):
    """
    Reloads the same question from the database, resetting the UI
    as if the question had not been answered yet.
    """
    return render_question_by_id(question_id)
    
@app.route("/question", methods=["GET"])
def question():
    """
    Selects a new question according to the VBA‑replicated logic
    and renders it cleanly.
    """
    with open_repo() as repo:
        q = repo.get_random_question(
            filtro_revisados=session_filtro_revisados,
            filtro_fallados=session_filtro_fallados,
        )

    return render_question_by_id(q.id)

@app.route("/next", methods=["POST"])
def next_question():
    """
    Persists the final answer and updates session counters.
    """
    global session_correct, session_incorrect
    global session_current_question_id, session_current_correct_option

    data = request.get_json()
    question_id = data["question_id"]
    selected = data["selected_option"]
    review_flag = data.get("review_flag", False)

    # ------------------------------
    # ✅ Validación de coherencia
    # ------------------------------
    if question_id != session_current_question_id:
        return ("Estado de sesión inconsistente", 400)

    # ------------------------------
    # ✅ Corrección con ORDEN BARJADO
    # ------------------------------
    is_correct = selected == session_current_correct_option

    # ------------------------------
    # ✅ Persistencia en BD
    # ------------------------------
    with open_repo() as repo:
        repo.mark_as_seen(question_id)
        repo.save_answer(
            AnswerResult(
                question_id=question_id,
                selected_option=selected,
                correct=is_correct
            )
        )
        repo.set_review_flag(question_id, review_flag)

    # ------------------------------
    # ✅ Contadores de sesión
    # ------------------------------
    if is_correct:
        session_correct += 1
    else:
        session_incorrect += 1

    return ("", 204)

@app.route("/start", methods=["POST"])
def start_exam():
    """
    Inicia una nueva sesión de examen (COMENZAR).

    Lee los filtros seleccionados en la pantalla inicial
    y reinicia los contadores de sesión.
    """
    global session_mode, session_initialized
    global session_filtro_revisados, session_filtro_fallados

    session_mode = "new"
    session_initialized = False

    # Leer filtros desde el formulario
    session_filtro_revisados = request.form.get("filtro_revisados", "Todos")
    session_filtro_fallados = request.form.get("filtro_fallados", "Todos")

    return redirect(url_for("question"))

@app.route("/continue", methods=["POST"])
def continue_exam():
    """
    Continúa una sesión anterior.

    Lee los filtros seleccionados en la pantalla inicial
    pero mantiene los contadores heredados.
    """
    global session_mode, session_initialized
    global session_filtro_revisados, session_filtro_fallados

    session_mode = "continue"
    session_initialized = False

    # Leer filtros desde el formulario
    session_filtro_revisados = request.form.get("filtro_revisados", "Todos")
    session_filtro_fallados = request.form.get("filtro_fallados", "Todos")

    return redirect(url_for("question"))

@app.route("/", methods=["GET"])
def home():
    """
    Página principal del sistema de exámenes.
    """
    return render_template_string(
        HOME_PAGE,
        app_name=app_meta.APP_NAME,
        app_version=app_meta.APP_VERSION,
        app_estado=app_meta.APP_ESTADO,
        app_build=app_meta.APP_BUILD,
    )

@app.route("/reset", methods=["POST"])
def reset_db():
    """
    Resetea completamente la base de datos de preguntas.

    Pone a cero las columnas:
    - VIS
    - COR
    - REV
    - OK
    - KO
    """
    with open_repo() as repo:
        repo.reset_statistics()

    return redirect(url_for("home"))
    
@app.route("/about")
def about():
    """
    Página 'Acerca de'.
    """
    return render_template_string(ABOUT_PAGE,
        app_name=app_meta.APP_NAME,
        app_version=app_meta.APP_VERSION,
        app_estado=app_meta.APP_ESTADO,
        app_build=app_meta.APP_BUILD,
    )

@app.route("/license")
def license():
    """
    Redirige a la licencia oficial GPLv3.
    """
    return redirect("https://www.gnu.org/licenses/gpl-3.0.html")

@app.route("/instructions")
def instructions():
    """
    Página de instrucciones / manual de usuario.
    """
    return render_template_string(INSTRUCTIONS_PAGE)

@app.route("/question_db/<question_id>", methods=["GET", "POST"])
def question_db(question_id):
    with open_repo() as repo:

        if request.method == "POST":
            data = {}

            for field, raw_value in request.form.items():
                if field in ("VIS", "COR", "REV", "OK", "KO"):
                    try:
                        value = int(raw_value)
                    except (TypeError, ValueError):
                        value = 0
                else:
                    value = raw_value

                data[field] = value

            repo.update_question_raw(question_id, data)

            return (
                "<h2>✅ Cambios guardados</h2>"
                "<button onclick='window.close()'>Cerrar</button>"
            )

        data = repo.get_question_raw(question_id)

    return render_template_string(QUESTION_DB_PAGE, data=data)

@app.route("/data_view/<name>")
def data_view(name):
    """
    Displays auxiliary data views:
    - Excel: worksheet rendered as HTML table
    - Access: table rendered as HTML table
    """
    #ext = DATA_FILE.suffix.lower()
    ext = Path(DATA_FILE).suffix.lower()
    name_escaped = escape(name)

    # =====================================================
    # ✅ EXCEL: leer hoja
    # =====================================================
    if ext in (".xls", ".xlsx", ".xlsm"):
        wb = load_workbook(DATA_FILE, data_only=True)

        if name not in wb.sheetnames:
            return f"<h2>La hoja '{name_escaped}' no existe</h2>"

        sheet = wb[name]

        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

    # =====================================================
    # ✅ ACCESS: leer tabla
    # =====================================================
    elif ext in (".mdb", ".accdb", ".db"):
        with open_repo() as repo:
            cursor = getattr(repo, "_cursor", None) or getattr(repo, "cursor", None)

            if cursor is None:
                return "<h2>No se pudo acceder al cursor de datos</h2>"

            # ------------------------------
            # Obtener lista de tablas
            # ------------------------------
            tables = []

            import sqlite3
            if isinstance(cursor, sqlite3.Cursor):
                # ✅ SQLite
                res = cursor.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
                tables = [r[0].upper() for r in res]
            else:
                # ✅ Access (pyodbc)
                for r in cursor.tables(tableType="TABLE"):
                    tables.append(r.table_name.upper())

            if name.upper() not in tables:
                return f"<h2>La tabla '{escape(name)}' no existe</h2>"

            # ------------------------------
            # Leer tabla completa
            # ------------------------------
            result = cursor.execute(f'SELECT * FROM "{name}"')
            columns = [col[0] for col in cursor.description]
            rows = result.fetchall()

            headers = columns
            data_rows = rows
    else:
        return "<h2>Tipo de base de datos no soportado</h2>"

    # =====================================================
    # ✅ Renderizado HTML común
    # =====================================================
    html = f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>{name_escaped}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                padding: 10px;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                border: 1px solid #999;
                padding: 4px 6px;
                font-size: 13px;
            }}
            th {{
                background: #e6f6ff;
                position: sticky;
                top: 0;
            }}
        </style>
    </head>
    <body>
        <h2>{name_escaped}</h2>
        <table>
            <tr>
    """

    for h in headers:
        html += f"<th>{escape(str(h))}</th>"

    html += "</tr>"

    for row in data_rows:
        html += "<tr>"
        for cell in row:
            val = "" if cell is None else escape(str(cell))
            html += f"<td>{val}</td>"
        html += "</tr>"

    html += """
        </table>
    </body>
    </html>
    """

    return html
    
def run_server(
    data_file,
    port,
    protocol,
    cert_file="",
    key_file="",
):
    global DATA_FILE, http_server

    DATA_FILE = str(data_file)

    ssl_context = None
    if protocol == "HTTPS" and cert_file and key_file:
        ssl_context = (cert_file, key_file)

    # ✅ Crear servidor WSGI controlable
    http_server = make_server(
        "127.0.0.1",
        port,
        app,
        ssl_context=ssl_context,
    )

    print(f"🟢 Servidor Flask arrancado en puerto {port}")
    http_server.serve_forever()

def stop_flask_server():
    global http_server

    if http_server:
        print("🟥 Deteniendo servidor Flask...")
        http_server.shutdown()
        http_server.server_close()
        http_server = None
        print("🔴 Servidor Flask detenido")