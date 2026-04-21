# This file is part of Servidor de Exámenes Exam-Web by Naidel.
#
# Copyright (C) 2024–2026 by Naidel
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
# See the GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""
Excel-based implementation of the QuestionRepository interface.

This module provides a safe and robust adapter for using an Excel (.xlsm)
file as the persistent storage of exam questions and statistics.

Macros and VBA code are preserved using openpyxl with keep_vba=True.
"""

import random
import time
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from data.base import (
    AnswerOption,
    AnswerResult,
    DataRepositoryLockedError,
    DataRepositoryNotAvailableError,
    Question,
    QuestionRepository,
)

_rng = random.Random()

class ExcelQuestionRepository(QuestionRepository):
    """
    Repository that stores and retrieves exam questions from an Excel file.
    """

    def __init__(self, file_path: str | Path, sheet_name: Optional[str] = None):
        """
        Initializes the repository.

        Args:
            file_path: Path to the Excel file (.xlsm).
            sheet_name: Name of the worksheet containing the questions.
        """
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
        self._workbook = None
        self._sheet = None
        self._columns: dict[str, int] = {}

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()
    
    def open(self) -> None:
        """
        Opens the Excel workbook and prepares column indexing.

        Raises:
            DataRepositoryNotAvailableError: If the file does not exist.
            DataRepositoryLockedError: If the file cannot be opened.
        """
        if not self.file_path.exists():
            raise DataRepositoryNotAvailableError(
                f"Excel file not found: {self.file_path}"
            )

        try:
            self._workbook = load_workbook(
                self.file_path,
                keep_vba=True,
                data_only=False,
            )
        except Exception as exc:
            raise DataRepositoryLockedError(
                "Excel file is locked or cannot be opened"
            ) from exc

        self._sheet = (
            self._workbook[self.sheet_name]
            if self.sheet_name
            else self._workbook.active
        )

        header = next(self._sheet.iter_rows(min_row=1, max_row=1))
        self._columns = {
            str(cell.value).strip().upper(): index
            for index, cell in enumerate(header)
            if cell.value is not None
        }

    def close(self) -> None:
        """
        Saves and closes the Excel workbook.
        """
        if self._workbook is not None:
            self._workbook.save(self.file_path)
            self._workbook.close()
            self._workbook = None
            self._sheet = None

    def get_all_questions(self) -> Iterable[Question]:
        """
        Returns all questions stored in the Excel file.

        Yields:
            Question instances.
        """
        for row in self._sheet.iter_rows(min_row=2):
            yield self._row_to_question(row)

    def get_random_question(
        self,
        filtro_revisados: str = "Todos",
        filtro_fallados: str = "Todos",
    ):
        """
        Returns a random question following the exact VBA logic:

        Priority order:
        1. Absolute priority to questions marked for review (REV = 1)
        2. Secondary priority to failed questions (VIS = 1 and COR = 0)
        3. Apply filters:
           - Revisados / No revisados
           - Fallados / Todos
        4. Exclude hidden rows
        5. Random selection with limited attempts
        """

        sheet = self._sheet
        col = self._columns
        
        # ✅ Reseed como hace Excel internamente
        _rng.seed(time.time_ns())

        total_questions = sheet.max_row - 1
        max_attempts = total_questions * 2

        last_candidate = None

        for _ in range(max_attempts):

            # Selección aleatoria de fila
            row_index = _rng.randint(2, sheet.max_row)
            
            row = sheet[row_index]

            # Fila oculta → excluir siempre
            if sheet.row_dimensions[row_index].hidden:
                continue

            vis = row[col["VIS"]].value or 0
            cor = row[col["COR"]].value or 0
            rev = row[col["REV"]].value or 0

            # PRIORIDAD ABSOLUTA: revisadas
            if rev == 1:
                return self._row_to_question(row)

            # Inicialmente permitimos
            allow_no_revisados = True
            allow_fallados = True

            # Filtro: solo no revisados
            if filtro_revisados not in ("Todos", "", None):
                if rev == 1:
                    allow_no_revisados = False

            # Filtro: solo fallados
            if filtro_fallados not in ("Todos", "", None):
                if cor == 1:
                    allow_fallados = False
                else:
                    # Prioridad implícita a falladas
                    allow_no_revisados = True

            if allow_no_revisados and allow_fallados:
                last_candidate = row
                break

        # Si no se encontró perfecta, devolver última candidata válida
        if last_candidate is not None:
            return self._row_to_question(last_candidate)

        # Caso extremo: devolver cualquiera no oculta
        for row in sheet.iter_rows(min_row=2):
            if not sheet.row_dimensions[row[0].row].hidden:
                return self._row_to_question(row)

        raise RuntimeError("No valid question found")

    def mark_as_seen(self, question_id: str) -> None:
        """
        Marks a question as seen (VIS = 1).

        Args:
            question_id: Identifier of the question.
        """
        for row in self._sheet.iter_rows(min_row=2):
            if self._value(row, "NOMBRE") == question_id:
                self._set(row, "VIS", 1)
                return

    def save_answer(self, result: AnswerResult) -> None:
        """
        Persists the final answer of a question.

        COR reflects only the last attempt.
        OK and KO are cumulative counters.

        Args:
            result: Final answer result selected by the user.
        """
        for row in self._sheet.iter_rows(min_row=2):
            if self._value(row, "NOMBRE") == result.question_id:
                if result.correct:
                    self._increment(row, "OK")
                    self._set(row, "COR", 1)
                else:
                    self._increment(row, "KO")
                    self._set(row, "COR", 0)
                return

    def get_raw_status(self):
        """
        Returns the raw VIS and COR status for all questions.

        Yields:
            tuple(int, int): (VIS, COR) values for each question.
        """
        for row in self._sheet.iter_rows(min_row=2):
            vis = row[self._columns["VIS"]].value
            cor = row[self._columns["COR"]].value
            yield vis, cor
            
    def reset_statistics(self) -> None:
        """
        Resets VIS, COR, REV, OK and KO for all questions.
        """
        for row in self._sheet.iter_rows(min_row=2):
            self._set(row, "VIS", 0)
            self._set(row, "COR", 0)
            self._set(row, "REV", 0)
            self._set(row, "OK", 0)
            self._set(row, "KO", 0)

    def _row_to_question(self, row) -> Question:
        """
        Converts a worksheet row into a Question object.

        Args:
            row: Worksheet row.

        Returns:
            Question instance.
        """
        options = [
            AnswerOption("A", self._value(row, "A"), self._value(row, "RA")),
            AnswerOption("B", self._value(row, "B"), self._value(row, "RB")),
            AnswerOption("C", self._value(row, "C"), self._value(row, "RC")),
            AnswerOption("D", self._value(row, "D"), self._value(row, "RD")),
        ]

        return Question(
            id=str(self._value(row, "NOMBRE")),
            statement=str(self._value(row, "PREGUNTA")),
            options=options,
            correct_option=str(self._value(row, "R")),
            topic=str(self._value(row, "UN")),
            question_type=str(self._value(row, "TIPO")),
            study_notes=self._value(row, "ESTUDIO"),
            flagged=bool(self._value(row, "REV")),
        )

    def set_review_flag(self, question_id: str, flagged: bool) -> None:
            """
            Sets or clears the review flag (REV) for a question.

            This method does NOT depend on answers. It only modifies
            the REV column for the specified question.

            Args:
                question_id: Identifier of the question (column NOMBRE).
                flagged: True to set REV = 1, False to set REV = 0.
            """
            for row in self._sheet.iter_rows(min_row=2):
                if row[self._columns["NOMBRE"]].value == question_id:
                    row[self._columns["REV"]].value = 1 if flagged else 0
                    return

    def _value(self, row, column: str):
        return row[self._columns[column]].value

    def _set(self, row, column: str, value) -> None:
        row[self._columns[column]].value = value

    def _increment(self, row, column: str) -> None:
        current = self._value(row, column) or 0
        self._set(row, column, current + 1)
        
    def calculate_global_stats(self):
        total_correct = 0
        total_incorrect = 0
        total_questions = 0
        seen_questions = 0

        for row in self._sheet.iter_rows(min_row=2):
            total_questions += 1

            ok = int(row[self._columns["OK"]].value or 0)
            ko = int(row[self._columns["KO"]].value or 0)
            vis = int(row[self._columns["VIS"]].value or 0)

            total_correct += ok
            total_incorrect += ko
            if vis == 1:
                seen_questions += 1

        total_attempts = total_correct + total_incorrect

        return {
            "correct": total_correct,
            "incorrect": total_incorrect,
            "percentage": round(
                (total_correct / total_attempts) * 100, 2
            ) if total_attempts > 0 else 0,
            "total_questions": total_questions,
            "seen_questions": seen_questions,
            "seen_percentage": round(
                (seen_questions / total_questions) * 100, 2
            ) if total_questions > 0 else 0,
        }
        
    def get_question_detail(self, question_id: str):
        for row in self._sheet.iter_rows(min_row=2):
            if row[self._columns["NOMBRE"]].value == question_id:
                return {
                    "question": self._row_to_question(row),
                    "vis": int(row[self._columns["VIS"]].value or 0),
                    "cor": int(row[self._columns["COR"]].value or 0),
                    "rev": int(row[self._columns["REV"]].value or 0) == 1,
                    "ok": int(row[self._columns["OK"]].value or 0),
                    "ko": int(row[self._columns["KO"]].value or 0),
                }
        return None
        
    def get_question_raw(self, question_id: str) -> dict:
        for row in self._sheet.iter_rows(min_row=2):
            if row[self._columns["NOMBRE"]].value == question_id:
                data = {}
                for col, idx in self._columns.items():
                    value = row[idx].value
                    data[col] = "" if value is None else value
                return data
        raise ValueError("Pregunta no encontrada")


    def update_question_raw(self, question_id: str, data: dict) -> None:
        for row in self._sheet.iter_rows(min_row=2):
            if row[self._columns["NOMBRE"]].value == question_id:
                for field, value in data.items():
                    if field in self._columns:
                        row[self._columns[field]].value = value
                return
        raise ValueError("Pregunta no encontrada")        